import os, time, datetime, string, random
from openpyxl import Workbook, load_workbook, utils
from openpyxl.styles import Alignment, Border, Side, PatternFill


class ExcelInteract:
    __wb = None
    __ws = None
    __totalBalance = 0
    __totalDebtBalance = 0
    __rowLimit = 999999

    # thick border style
    __thickBorder = Border(
        top=Side(style='thick'),
        left=Side(style='thick'),
        right=Side(style='thick'),
        bottom=Side(style='thick')
    )
    # thin border style
    __thinBorder = Border(
            top=Side(style='thin'),
            left=Side(style='thin'),
            right=Side(style='thin'),
            bottom=Side(style='thin')
        )
    # colors
    __colors = {
        'green':PatternFill(start_color='FF19E012', end_color='FF19E012', fill_type='solid'),
        'red':PatternFill(start_color='e01212', end_color='e01212', fill_type='solid'),
        'orange':PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
    }


    # initialization code
    def __init__(self):
        print('Initializing workbook .....')
        # get file path for the workbook
        self.__excelFilePath = os.path.join(os.getcwd(), 'self-finance.xlsx')

        # create workbook
        if not os.path.exists(self.__excelFilePath):
            self.__wb = Workbook()
            self.__ws = self.__wb.active
            self.__ws.title = 'wallet1'

            self.saveWB()
        # load existing workbook
        else:
            self.__wb = load_workbook('self-finance.xlsx')
            self.__ws = self.__wb.active

    # to save xlsx file changes
    def saveWB(self):
        # format file style first before saving changes
        self.__formatWB()

        while True:
            try:
                self.__wb.save(os.path.join(os.getcwd(), 'self-finance.xlsx'))
                break
            except Exception as e:
                print(f"[error]: {e}\n[system]: Close the finance file to proceed, retrying in 5s")
                time.sleep(5)

    # apply formatting to xlsx file
    def __formatWB(self):
        # styling transactions column
        self.__ws['A1'] = 'Transactions'
        self.__ws['A2'] = 'Date'
        self.__ws['B2'] = 'Amount'
        self.__ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=2)
        
        # styling debts column
        self.__ws['D1'] = 'Debts'
        self.__ws['D2'] = 'Date'
        self.__ws['E2'] = 'Amount'
        self.__ws.merge_cells(start_row=1, end_row=1, start_column=4, end_column=5)
        
        # styling total column
        self.__ws['G1'] = 'Total Balance: '
        self.__ws['G2'] = 'Total Debt: '
        for hColumnRow in range(1, 3): # row 1 and 2, column 8 (numerical values)
            cell = self.__ws.cell(row=hColumnRow, column=8)
            cell.alignment = Alignment(horizontal='right', vertical='center')
        
        # auto fit columns
        for column in range(1, 9): # columns A to H
            self.__ws.column_dimensions[utils.get_column_letter(column)].auto_size = True
        
        # set text in cells alignments
        for c in range(1, 6): # columns A to E
            for r in range(1, self.__ws.max_row+1): # rows 1 to max
                cell = self.__ws.cell(column=c, row=r)
                # if the cell is in amount column set text to right
                if (c == 2 or c == 5) and (r > 2):
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                else: # if not, then center
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # setting borders
        for borderC in range(1, self.__ws.max_column+1):
            for borderR in range(1, self.__ws.max_row+1):
                cell = self.__ws.cell(column=borderC, row=borderR)

                if not (borderC == 3 or borderC == 6):
                    # set thick border to Transaction, Debt, Totals Cells
                    if (cell.row < 3) or (cell.row < 3 and cell.column < 9):
                        cell.border = self.__thickBorder
                    else: # thin border to else
                        if borderC < 7 and self.__ws.cell(column=borderC, row=borderR).value != None:
                            cell.border = self.__thinBorder

    # used to access transaction column
    def __accessTransactions(self, amount, mode):
        # getting amount cell's color representation
        cell_color = None
        if mode == 'deposit':
            cell_color = self.__colors['green']
        else:
            cell_color = self.__colors['red']

        # iterate over transaction column
        for row_index in range(3, self.__rowLimit):
            date_cell = self.__ws.cell(column=1, row=row_index)
            if date_cell.value == None:
                amount_cell = self.__ws.cell(column=2, row=row_index)
                date_cell.value = datetime.datetime.now().strftime('%x')
                amount_cell.value = f"{amount:,.2f}"
                amount_cell.fill = cell_color
                break
    
    # used to access debt column
    def __accessDebts(self, amount, mode):
        cell_color = self.__colors['orange']
        debt_count = 0

        for row_index in range(3, self.__rowLimit):
            date_cell = self.__ws.cell(column=4, row=row_index)
            amount_cell = self.__ws.cell(column=5, row=row_index)

            if (date_cell.value == None) and (mode == "add"):
                date_cell.value = datetime.datetime.now().strftime('%x')
                amount_cell.value = f"{amount:,.2f}"
                amount_cell.fill = cell_color
                break
            elif mode == "pay":
                if date_cell.value != None:
                    print(f"\t{row_index-2}. {date_cell.value} | {amount_cell.value}")
                    debt_count += 1
                else:
                    break
        
        if mode == "pay":
            print(f"\nChoose debt number to pay [1 - {debt_count}]\n")
            
            # user gets to choose which debt to pay/remove
            cell_row_choice = Interface.userInput(1, debt_count) + 2

            date_cell = self.__ws.cell(column=4, row=cell_row_choice)
            amount_cell = self.__ws.cell(column=5, row=cell_row_choice)
            total_balance_value = float(self.__ws['H1'].value.replace(',', ''))

            # if the debt trying to pay is greater than the total balance, cannot proceed
            if float(amount_cell.value.replace(',', '')) > total_balance_value:
                print("[error]: you're trying to pay a debt that is larger than your total balance, process terminated")
                return None
            else:
                self.withdraw(True, amount_cell.value)

            # setting the date and amount to nothing
            date_cell.value = None
            amount_cell.value = None
            amount_cell.fill = PatternFill(fill_type='none')
            
            # rearrange debts to fill the deleted debt
            self.__sort_debt_from_deletedRow(cell_row_choice)
    
    # To move all the existing values below of the deleted debt row upwards
    def __sort_debt_from_deletedRow(self, row):

        while True:
            # gets current position of the deleted debt(date & amount)
            current_date_cell = self.__ws.cell(column=4, row=row)
            current_amount_cell = self.__ws.cell(column=5, row=row)

            # cells below the current
            date_cell_below = self.__ws.cell(column=4, row=row + 1)
            amount_cell_below = self.__ws.cell(column=5, row=row + 1)

            if date_cell_below.value != None and amount_cell_below.value != None:
                current_date_cell.value = date_cell_below.value # copying the date
                current_amount_cell.value = amount_cell_below.value # copying the amount
                current_amount_cell.fill = self.__colors['orange'] # copying the fill

                # removing everything from the cells below the current
                date_cell_below.value = None
                amount_cell_below.value = None
                amount_cell_below.fill = PatternFill(fill_type='none')

                row += 1 # increment for next row
            else:
                break # breaks if it reaches the end of the debt column

    # set the total balance
    def _getTotalBalance(self):
        # iterate through the amount column
        for row_index in range(2, self.__ws.max_row):
            cell = self.__ws.cell(column=2, row=row_index+1)
            cell_color = cell.fill.start_color.rgb # get cell color fill
            cell_value = 0

            if cell.value != None:
                cell_value = float(cell.value.replace(',', '')) # get float value
                
            if cell_color == 'FF19E012':
                self.__totalBalance += cell_value
            else:
                self.__totalBalance -= cell_value
        
        total_cell = self.__ws['H1']
        total_cell.value = f"{self.__totalBalance:,.2f}" # apply total in xlsx file
        self.__totalBalance = 0 # reset total balance in object
    
    # set the total debt balance
    def _getTotalDebtBalance(self):
        
        for row_index in range(2, self.__ws.max_row):

            cell = self.__ws.cell(column=5, row=row_index+1)

            if cell.value != None:
                cell_value = float(cell.value.replace(',', '')) # get float value
            else:
                break

            self.__totalDebtBalance += cell_value
        
        debt_total_cell = self.__ws['H2']
        debt_total_cell.value = f"{self.__totalDebtBalance:,.2f}"
        self.__totalDebtBalance = 0

    def deposit(self):
        deposit = Interface.validateNumInput("deposit")
        # put the deposit in the amount w/ date
        self.__accessTransactions(deposit, 'deposit')

    def withdraw(self, for_debt = False, debt_amount = 0):
        cell_total = self.__ws['H1'].value # get total from xlsx file
        withdraw = None

        if for_debt:
            withdraw = float(debt_amount.replace(',', ''))
        else:
            # validate if the withdraw amount is acceptable
            withdraw = Interface.validateNumInput("withdraw", float(cell_total.replace(',', '')), True) # compare if withdrawal is greater than total, therefore cannot withdraw
        
        # put the withdraw in the amount w/ date
        self.__accessTransactions(withdraw, 'withdraw')

    def debts(self, mode):
        if mode == 1:
            # get debt
            debt = Interface.validateNumInput("debt")
            self.__accessDebts(debt, "add")
        else:
            # get total debt to compare if a debt exists or not
            debt_total = float(self.__ws['H2'].value.replace(',', ''))

            if debt_total > 0:
                self.__accessDebts(0, "pay")
            else:
                print(f"[system]: No debt to pay, process terminated")

    def view_totals(self):
        total_balance = self.__ws['H1'].value
        total_debt_balance = self.__ws['H2'].value

        # prints all the total balances
        print(f"\nTotal balance: {total_balance}\nTotal Debt balance: {total_debt_balance}\n")
    
    def delete_all_transactions(self):

        if Interface.validateDeletion():
            # deletes the rows starting from 3
            self.__ws.delete_rows(3, self.__ws.max_row)
            print(f"\n[system]: input valid, all transactions deleted")
        else:
            print(f"\n[system]: incorrect input, process terminated")

class Interface:
    
    # choice input checker
    def userInput(min = 0, max = 5) -> int:
        currentInput = None

        # loops back whenever the input is invalid
        while True:
            try:
                currentInput = int(input("input> "))

                # if input is within the set range, accept it
                if (currentInput < max+1 and currentInput > min-1): 
                        break
                # retry
                else:
                    print(f"{currentInput} is invalid")
            # retry
            except:
                print(f"error: invalid input, can only accept a number(int)")
        
        # return valid input as integer
        return int(currentInput)
    

    # number input checker
    def validateNumInput(mode_if, limit = 1, withdrawal = False) -> float:
        currentInput = None # input holder

        # loops back until the value input is satisfied
        while True:
            try:                                # mode_if is the message guide if for withdraw or deposit or debt (ex. deposit-amount>)
                currentInput = abs(float(input(f"{mode_if}-amount> ")))
            except:
                print(f"error: invalid input, can only accept a number(float)")
            else:
                # if for withdrawal and the input is greater than total balance, deny it
                if withdrawal and (currentInput > limit):
                    print(f"error: cannot withdraw, {currentInput} is greater than your total balance")
                # if not for withdrawal and the input is less than the limit, deny it
                elif not withdrawal and (currentInput < limit):
                    print(f"{currentInput} is invalid")
                # break free from loop, if neither
                else:
                    break
        
        # return input as float
        return float(currentInput)

    # prompts the user to type the exact string given
    def validateDeletion() -> bool:
        letters = string.ascii_letters
        nums = string.digits
        string_guide = ''

        # append a random letter or digit to the guide
        for i in range(10):
            string_guide += random.choice([random.choice(letters), random.choice(nums)])
        
        str_input = input(f'Retype "{string_guide}" to confirm deletion: ')

        # if the input is equal to the guide, confirm deletion(return True)
        if str_input == string_guide:
            return True
        else:
            return False
            

def main():
    # initialize class/object
    account = ExcelInteract()

    # MAIN loop of the program
    while True:
        account._getTotalBalance() # update total balance
        account._getTotalDebtBalance() # update total debt balance

        print('\n-- Chiharu\'s Finance System(CFS) --')
        print("\t1. Deposit\n\t2. Withdraw\n\t3. Debts\n\t4. View total balances\n\t5. Exit\n")
        user_input = Interface.userInput() # get proper input

        match (user_input):
            case 0:
                account.delete_all_transactions()
            case 1:
                account.deposit()
            case 2:
                account.withdraw()
            case 3:
                print("\t1. Add debt\n\t2. Pay debt\n")
                debt_mode = Interface.userInput(1, 2) # user selects if to add or pay/remove debt
                account.debts(debt_mode)
            case 4:
                account.view_totals()
            case _:
                print("\nExiting . . . . .")
                account.saveWB()

                print('[system]: Changes saved! Open "self-finance" excel file for better data representation\n')
                break

if __name__ == "__main__":
    main()